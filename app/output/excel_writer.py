"""Generates enriched Excel workbook with cross-referenced data."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import EnrichedItem, GapItem, GapSeverity, MatchStatus


# Color scheme
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_SECTION = Font(bold=True, size=11)
FONT_NORMAL = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_FILL = {
    MatchStatus.MATCHED: FILL_GREEN,
    MatchStatus.PARTIAL: FILL_YELLOW,
    MatchStatus.UNMATCHED: FILL_RED,
    MatchStatus.GAP: FILL_RED,
}

SEVERITY_FILL = {
    GapSeverity.HIGH: FILL_RED,
    GapSeverity.MEDIUM: FILL_YELLOW,
    GapSeverity.LOW: FILL_GREEN,
}


class ExcelWriter:
    """Generates enriched Excel workbooks."""

    def write(
        self,
        enriched_items: list[EnrichedItem],
        gaps: list[GapItem],
        output_path: str | Path,
    ):
        wb = Workbook()

        # Group items by sheet name
        items_by_sheet: dict[str, list[EnrichedItem]] = {}
        for item in enriched_items:
            sheet = item.magnaskra_item.sheet_name
            items_by_sheet.setdefault(sheet, []).append(item)

        # Create a sheet for each original magnskrá sheet
        first = True
        for sheet_name, items in items_by_sheet.items():
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
            self._write_items_sheet(ws, items)

        # Create gaps sheet
        if gaps:
            ws_gaps = wb.create_sheet(title="Vantar í magnskrá")
            self._write_gaps_sheet(ws_gaps, gaps)

        # Create summary sheet at the beginning
        ws_summary = wb.create_sheet(title="Samantekt", index=0)
        self._write_summary_sheet(ws_summary, enriched_items, gaps)

        wb.save(output_path)

    def _write_items_sheet(self, ws, items: list[EnrichedItem]):
        """Write enriched items to a worksheet."""
        headers = [
            "Nr.",
            "Heiti verkþáttar",
            "Magn",
            "Eining",
            "Verklýsing kröfur",
            "BC vara (SKU)",
            "BC lýsing",
            "Listaverð",
            "Staða",
            "Athugasemdir",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data
        for row_idx, enriched in enumerate(items, 2):
            item = enriched.magnaskra_item

            # Section number
            ws.cell(row=row_idx, column=1, value=item.section_nr).font = (
                FONT_SECTION if item.is_header else FONT_NORMAL
            )

            # Description
            ws.cell(row=row_idx, column=2, value=item.description).font = (
                FONT_SECTION if item.is_header else FONT_NORMAL
            )

            # Quantity & unit
            if item.quantity is not None:
                ws.cell(row=row_idx, column=3, value=item.quantity)
            ws.cell(row=row_idx, column=4, value=item.unit or "")

            # Verklýsing requirements
            if enriched.verklysing_requirements:
                reqs_text = "\n".join(
                    f"[{r.category}] {r.text}"
                    for r in enriched.verklysing_requirements
                )
                cell = ws.cell(row=row_idx, column=5, value=reqs_text)
                cell.alignment = Alignment(wrap_text=True)

            # BC product match
            if enriched.bc_product:
                ws.cell(row=row_idx, column=6, value=enriched.bc_product.sku)
                ws.cell(row=row_idx, column=7, value=enriched.bc_product.description)
                if enriched.bc_product.unit_price:
                    cell = ws.cell(row=row_idx, column=8, value=enriched.bc_product.unit_price)
                    cell.number_format = '#,##0'

            # Status
            if not item.is_header:
                status_cell = ws.cell(row=row_idx, column=9, value=enriched.match_status.value)
                if enriched.match_status in STATUS_FILL:
                    status_cell.fill = STATUS_FILL[enriched.match_status]

            # Notes
            ws.cell(row=row_idx, column=10, value=enriched.notes)

            # Apply section header styling
            if item.is_header:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = FILL_SECTION
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER
            else:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).border = THIN_BORDER

        # Set column widths
        widths = [12, 45, 10, 8, 50, 18, 40, 12, 12, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

    def _write_gaps_sheet(self, ws, gaps: list[GapItem]):
        """Write gap analysis results."""
        headers = ["Kafli", "Titill", "Krafa", "Alvarleiki", "Tillaga"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.border = THIN_BORDER

        for row_idx, gap in enumerate(gaps, 2):
            ws.cell(row=row_idx, column=1, value=gap.source_section).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=gap.source_title).border = THIN_BORDER

            req_cell = ws.cell(row=row_idx, column=3, value=gap.requirement_text)
            req_cell.border = THIN_BORDER
            req_cell.alignment = Alignment(wrap_text=True)

            sev_cell = ws.cell(row=row_idx, column=4, value=gap.severity.value)
            sev_cell.border = THIN_BORDER
            if gap.severity in SEVERITY_FILL:
                sev_cell.fill = SEVERITY_FILL[gap.severity]

            action_cell = ws.cell(row=row_idx, column=5, value=gap.suggested_action)
            action_cell.border = THIN_BORDER
            action_cell.alignment = Alignment(wrap_text=True)

        widths = [12, 30, 60, 12, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    def _write_summary_sheet(
        self, ws, enriched_items: list[EnrichedItem], gaps: list[GapItem]
    ):
        """Write summary/overview sheet."""
        ws.cell(row=1, column=1, value="VV Offer Tool - Samantekt").font = Font(
            bold=True, size=14
        )

        line_items = [e for e in enriched_items if not e.magnaskra_item.is_header]
        matched = sum(1 for e in line_items if e.match_status == MatchStatus.MATCHED)
        partial = sum(1 for e in line_items if e.match_status == MatchStatus.PARTIAL)
        unmatched = sum(1 for e in line_items if e.match_status == MatchStatus.UNMATCHED)

        high_gaps = sum(1 for g in gaps if g.severity == GapSeverity.HIGH)
        med_gaps = sum(1 for g in gaps if g.severity == GapSeverity.MEDIUM)
        low_gaps = sum(1 for g in gaps if g.severity == GapSeverity.LOW)

        stats = [
            ("", ""),
            ("Heildarfjöldi liða", len(line_items)),
            ("Samsvörun fundin (verklýsing + BC)", matched),
            ("Hlutasamsvörun", partial),
            ("Engin samsvörun", unmatched),
            ("", ""),
            ("Vantar í magnskrá (samtals)", len(gaps)),
            ("  Alvarlegt (efni/búnaður)", high_gaps),
            ("  Miðlungs (prófanir/skoðanir)", med_gaps),
            ("  Lágt (gott að hafa)", low_gaps),
        ]

        for i, (label, value) in enumerate(stats, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=bool(label and not label.startswith(" ")))
            if value != "":
                ws.cell(row=i, column=2, value=value)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
