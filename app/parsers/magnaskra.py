"""Parser for tilboðsskrá/magnskrá Excel files (bill of quantities)."""

import re
from pathlib import Path

import openpyxl

from app.models.schemas import MagnaskraItem

# Pattern to detect section numbers like 3.1, 3.1.1, 3.1.1.1
SECTION_NR_PATTERN = re.compile(r"^\d+(\.\d+)*$")

# Known header columns that identify a magnskrá sheet
HEADER_SIGNATURES = {"NR.", "MAGN", "EINING", "HEITI VERKÞÁTTAR", "EININGARVERÐ", "HEILDARVERÐ"}


class MagnaskraParser:
    """Parses tilboðsskrá/magnskrá Excel files into structured items."""

    def parse(self, file_path: str | Path) -> list[MagnaskraItem]:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        all_items = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not self._is_magnaskra_sheet(ws):
                continue
            items = self._parse_sheet(ws, sheet_name)
            all_items.extend(items)

        return all_items

    def _is_magnaskra_sheet(self, ws) -> bool:
        """Check if a sheet looks like a magnskrá by scanning the first 10 rows for header signatures."""
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            values = {str(v).strip().upper() for v in row if v is not None}
            # Need at least 3 matching header columns
            if len(values & HEADER_SIGNATURES) >= 3:
                return True
        return False

    def _parse_sheet(self, ws, sheet_name: str) -> list[MagnaskraItem]:
        header_row = self._find_header_row(ws)
        if header_row is None:
            return []

        col_map = self._map_columns(ws, header_row)
        items = []
        current_parents: dict[int, str] = {}  # depth -> section_nr

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            nr_val = self._cell_value(row, col_map.get("nr"))
            desc_val = self._cell_value(row, col_map.get("description"))
            qty_val = self._cell_value(row, col_map.get("quantity"))
            unit_val = self._cell_value(row, col_map.get("unit"))

            if not nr_val and not desc_val:
                continue

            nr_str = str(nr_val).strip() if nr_val else ""

            # Skip formula references like =Tilboðsblað!A1
            if nr_str.startswith("="):
                continue

            # Determine if this is a section header or a line item
            is_header = False
            if nr_str and SECTION_NR_PATTERN.match(nr_str):
                depth = nr_str.count(".")
                has_quantity = qty_val is not None and str(qty_val).strip() != ""
                is_header = not has_quantity

                # Track parent sections
                current_parents[depth] = nr_str
                # Clear deeper parents
                for d in list(current_parents.keys()):
                    if d > depth:
                        del current_parents[d]

            # Find parent section
            parent = None
            if nr_str and SECTION_NR_PATTERN.match(nr_str):
                depth = nr_str.count(".")
                if depth > 0 and (depth - 1) in current_parents:
                    parent = current_parents[depth - 1]

            # Parse quantity
            quantity = None
            if qty_val is not None:
                try:
                    quantity = float(qty_val)
                except (ValueError, TypeError):
                    pass

            description = str(desc_val).strip() if desc_val else ""
            if not nr_str and not description:
                continue

            # Handle rows without a section number (continuation descriptions)
            if not nr_str and description:
                # These are sub-descriptions or material-only rows (marked "Verkkaupi" etc.)
                # Attach to the previous item's section
                if items:
                    nr_str = items[-1].section_nr + "_cont"

            item = MagnaskraItem(
                section_nr=nr_str,
                description=description,
                quantity=quantity,
                unit=str(unit_val).strip() if unit_val else None,
                sheet_name=sheet_name,
                is_header=is_header,
                parent_section=parent,
            )
            items.append(item)

        return items

    def _find_header_row(self, ws) -> int | None:
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            values = {str(v).strip().upper() for v in row if v is not None}
            if len(values & HEADER_SIGNATURES) >= 3:
                return row_idx
        return None

    def _map_columns(self, ws, header_row: int) -> dict[str, int]:
        """Map semantic column names to column indices."""
        col_map = {}
        for col_idx, cell in enumerate(ws[header_row]):
            val = str(cell.value).strip().upper() if cell.value else ""
            if val in ("NR.", "NR"):
                col_map["nr"] = col_idx
            elif "HEITI" in val or "VERKÞÁTTAR" in val:
                col_map["description"] = col_idx
            elif val == "MAGN":
                col_map["quantity"] = col_idx
            elif val == "EINING":
                col_map["unit"] = col_idx
            elif "EININGARVERÐ" in val:
                col_map["unit_price"] = col_idx
            elif "HEILDARVERÐ" in val:
                col_map["total_price"] = col_idx
        return col_map

    def _cell_value(self, row, col_idx: int | None):
        if col_idx is None or col_idx >= len(row):
            return None
        val = row[col_idx].value
        if val is None:
            return None
        s = str(val).strip()
        if s in ("", "None"):
            return None
        return val
