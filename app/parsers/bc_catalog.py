"""Parser for BC (Business Central) product catalog Excel exports."""

from pathlib import Path

import openpyxl

from app.models.schemas import BCProduct, BCProductType

# Expected column headers in BC export
BC_HEADERS = {"Gerð", "Nr.", "Magn", "Lýsing", "Mælieiningarkóði"}


class BCCatalogParser:
    """Parses BC (Business Central) Excel exports into product catalog."""

    def parse(self, file_path: str | Path) -> list[BCProduct]:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if self._is_bc_sheet(ws):
                return self._parse_sheet(ws)

        return []

    def _is_bc_sheet(self, ws) -> bool:
        """Check if sheet has BC export headers."""
        if ws.max_row < 2:
            return False
        first_row_vals = {str(c.value).strip() for c in ws[1] if c.value}
        return len(first_row_vals & BC_HEADERS) >= 3

    def _parse_sheet(self, ws) -> list[BCProduct]:
        col_map = self._map_columns(ws)
        products = []
        current_section = None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            type_val = self._cell_str(row, col_map.get("type"))
            sku_val = self._cell_str(row, col_map.get("sku"))
            desc_val = self._cell_str(row, col_map.get("description"))
            qty_val = self._cell_value(row, col_map.get("quantity"))
            unit_val = self._cell_str(row, col_map.get("unit"))
            price_val = self._cell_value(row, col_map.get("unit_price"))
            cost_val = self._cell_value(row, col_map.get("cost_price"))

            if not type_val and not desc_val:
                continue

            # Determine product type
            product_type = BCProductType.VARA
            if type_val:
                type_upper = type_val.upper()
                if "ATHUGASEMD" in type_upper:
                    product_type = BCProductType.ATHUGASEMD
                    current_section = desc_val
                    continue  # Don't add comment rows as products, but track section
                elif "FORÐI" in type_upper or "FORDI" in type_upper:
                    product_type = BCProductType.FORDI

            # Parse quantity
            quantity = 0.0
            if qty_val is not None:
                try:
                    quantity = float(qty_val)
                except (ValueError, TypeError):
                    pass

            # Parse prices
            unit_price = None
            if price_val is not None:
                try:
                    unit_price = float(price_val)
                except (ValueError, TypeError):
                    pass

            cost_price = None
            if cost_val is not None:
                try:
                    cost_price = float(cost_val)
                except (ValueError, TypeError):
                    pass

            product = BCProduct(
                sku=sku_val or "",
                description=desc_val or "",
                quantity=quantity,
                unit=unit_val or "STK",
                product_type=product_type,
                section_comment=current_section,
                unit_price=unit_price,
                cost_price=cost_price,
            )
            products.append(product)

        return products

    def _map_columns(self, ws) -> dict[str, int]:
        """Map semantic column names to indices from header row."""
        col_map = {}
        for col_idx, cell in enumerate(ws[1]):
            val = str(cell.value).strip() if cell.value else ""
            val_lower = val.lower()
            if val == "Gerð":
                col_map["type"] = col_idx
            elif val == "Nr.":
                col_map["sku"] = col_idx
            elif val == "Magn":
                col_map["quantity"] = col_idx
            elif val == "Lýsing":
                col_map["description"] = col_idx
            elif val == "Mælieiningarkóði":
                col_map["unit"] = col_idx
            elif "ein.verð" in val_lower or "einingarverð" in val_lower:
                col_map["unit_price"] = col_idx
            elif "kostn" in val_lower or "sgm" in val_lower:
                col_map["cost_price"] = col_idx
        return col_map

    def _cell_str(self, row, col_idx: int | None) -> str | None:
        if col_idx is None or col_idx >= len(row):
            return None
        val = row[col_idx].value
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def _cell_value(self, row, col_idx: int | None):
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx].value
